# -*- coding: utf-8 -*-
"""diskguardian/services/report_service.py
PDF / CSV / Excel / JSON report generation.
"""

from __future__ import annotations
import io
import csv
import json
import datetime
from typing import Any

# ReportLab
try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, PageBreak
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    _RL = True
except ImportError:
    _RL = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _XLSX = True
except ImportError:
    _XLSX = False


# ─────────────────────────────────────────────────────────────────────────────
#  PDF Report
# ─────────────────────────────────────────────────────────────────────────────
def generate_pdf(scan: dict, user: dict, health: dict, risk: dict, benchmark: dict | None = None) -> bytes:
    """Generate a professional PDF report. Returns raw bytes."""
    if not _RL:
        raise RuntimeError("reportlab is not installed.")

    buf    = io.BytesIO()
    styles = getSampleStyleSheet()
    doc    = SimpleDocTemplate(buf, pagesize=A4,
                               leftMargin=2*cm, rightMargin=2*cm,
                               topMargin=2*cm, bottomMargin=2*cm)

    # Custom styles
    title_style = ParagraphStyle("ReportTitle",
        parent=styles["Heading1"], fontSize=22, textColor=colors.HexColor("#1e40af"),
        spaceAfter=4, alignment=TA_CENTER)
    sub_style = ParagraphStyle("SubTitle",
        parent=styles["Normal"], fontSize=10, textColor=colors.grey,
        alignment=TA_CENTER, spaceAfter=14)
    h2_style  = ParagraphStyle("H2",
        parent=styles["Heading2"], fontSize=13, textColor=colors.HexColor("#1e40af"),
        spaceBefore=14, spaceAfter=6)
    body_style = styles["BodyText"]
    body_style.fontSize = 10
    score_color = {
        "Excellent": colors.HexColor("#16a34a"),
        "Good":      colors.HexColor("#2563eb"),
        "Warning":   colors.HexColor("#d97706"),
        "Critical":  colors.HexColor("#dc2626"),
    }

    elements = []
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

    # ── Cover page ────────────────────────────────────────────────────────────
    elements.append(Spacer(1, 2*cm))
    elements.append(Paragraph("Hard Disk Analyzer", title_style))
    elements.append(Paragraph("Hard Drive Health Report", sub_style))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1e40af")))
    elements.append(Spacer(1, 0.5*cm))

    meta = [
        ["Generated:", now],
        ["User:", user.get("username", user.get("email", "—"))],
        ["Drive:", scan.get("model", scan.get("drive", "—"))],
        ["Serial:", scan.get("serial", "—")],
        ["Drive Type:", scan.get("drive_type", "—")],
    ]
    meta_tbl = Table(meta, colWidths=[4*cm, 12*cm])
    meta_tbl.setStyle(TableStyle([
        ("FONTSIZE",   (0, 0), (-1, -1), 10),
        ("FONTNAME",   (0, 0), (0, -1), "Helvetica-Bold"),
        ("TEXTCOLOR",  (0, 0), (0, -1), colors.HexColor("#1e40af")),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(meta_tbl)
    elements.append(Spacer(1, 1*cm))

    # ── Health Score ──────────────────────────────────────────────────────────
    elements.append(Paragraph("Health Score", h2_style))
    grade = health.get("grade", "Unknown")
    sc    = health.get("score", 0)
    hs_color = score_color.get(grade, colors.grey)
    score_tbl = Table([[f"Overall Score: {sc}/100", f"Grade: {grade}"]],
                       colWidths=[8*cm, 8*cm])
    score_tbl.setStyle(TableStyle([
        ("FONTSIZE",   (0, 0), (-1, -1), 14),
        ("FONTNAME",   (0, 0), (-1, -1), "Helvetica-Bold"),
        ("TEXTCOLOR",  (0, 0), (-1, -1), hs_color),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("BOX",        (0, 0), (-1, -1), 1, hs_color),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f0f9ff")),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(score_tbl)
    elements.append(Spacer(1, 0.5*cm))

    # ── Risk Prediction ───────────────────────────────────────────────────────
    elements.append(Paragraph("Risk Prediction", h2_style))
    risk_level = risk.get("label", "Unknown")
    elements.append(Paragraph(f"<b>Failure Risk:</b> {risk_level}", body_style))
    elements.append(Paragraph(f"<i>{risk.get('disclaimer', '')}</i>", body_style))
    for expl in risk.get("explanations", []):
        elements.append(Paragraph(f"• {expl}", body_style))
    elements.append(Spacer(1, 0.5*cm))

    # ── SMART Attributes ──────────────────────────────────────────────────────
    elements.append(Paragraph("SMART Attributes", h2_style))
    smart_attrs = scan.get("smart", {}).get("attributes", {})
    if smart_attrs:
        rows = [["ID", "Attribute", "Value", "Raw"]]
        for attr_id, attr in smart_attrs.items():
            rows.append([
                attr_id,
                attr.get("name", ""),
                str(attr.get("value", "")),
                str(attr.get("raw", "")),
            ])
        attr_tbl = Table(rows, colWidths=[2*cm, 8*cm, 3*cm, 3*cm])
        attr_tbl.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#1e40af")),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("GRID",        (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
            ("TOPPADDING",  (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(attr_tbl)
    else:
        elements.append(Paragraph("No SMART attributes available.", body_style))
    elements.append(Spacer(1, 0.5*cm))

    # ── Benchmark ─────────────────────────────────────────────────────────────
    if benchmark and not benchmark.get("error"):
        elements.append(Paragraph("Speed Benchmark", h2_style))
        bench_data = [
            ["Test",              "Speed"],
            ["Sequential Read",  f"{benchmark.get('seq_read_mbps',  0)} MB/s"],
            ["Sequential Write", f"{benchmark.get('seq_write_mbps', 0)} MB/s"],
            ["Random Read",      f"{benchmark.get('rand_read_mbps', 0)} MB/s"],
            ["Random Write",     f"{benchmark.get('rand_write_mbps',0)} MB/s"],
            ["Overall Score",    f"{benchmark.get('performance_score', 0)}/100 — {benchmark.get('rating', '')}"],
        ]
        bench_tbl = Table(bench_data, colWidths=[8*cm, 8*cm])
        bench_tbl.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#1e40af")),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 10),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("GRID",        (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
            ("TOPPADDING",  (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(bench_tbl)
        elements.append(Spacer(1, 0.5*cm))

    # ── Recommendations ───────────────────────────────────────────────────────
    checklist = health.get("checklist", [])
    if checklist:
        elements.append(Paragraph("AI Recommendations", h2_style))
        for item in checklist:
            elements.append(Paragraph(f"{item}", body_style))

    # ── Footer ────────────────────────────────────────────────────────────────
    elements.append(Spacer(1, 1*cm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
    elements.append(Paragraph(
        f"Generated by Hard Disk Analyzer v2.0.0 — {now}",
        ParagraphStyle("Footer", parent=body_style, fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
    ))

    doc.build(elements)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
#  CSV Export
# ─────────────────────────────────────────────────────────────────────────────
def generate_csv(scan: dict, health: dict) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Field", "Value"])
    writer.writerow(["Drive",         scan.get("drive", "")])
    writer.writerow(["Model",         scan.get("model", "")])
    writer.writerow(["Serial",        scan.get("serial", "")])
    writer.writerow(["Type",          scan.get("drive_type", "")])
    writer.writerow(["Health Score",  health.get("score", "")])
    writer.writerow(["Grade",         health.get("grade", "")])
    writer.writerow(["Temperature",   scan.get("smart", {}).get("temperature", "")])
    writer.writerow(["Power-On Hours", scan.get("smart", {}).get("power_on_hours", "")])
    writer.writerow([])
    writer.writerow(["SMART Attributes"])
    writer.writerow(["ID", "Name", "Value", "Raw"])
    for attr_id, attr in scan.get("smart", {}).get("attributes", {}).items():
        writer.writerow([attr_id, attr.get("name", ""), attr.get("value", ""), attr.get("raw", "")])
    return buf.getvalue().encode("utf-8")


# ─────────────────────────────────────────────────────────────────────────────
#  Excel Export
# ─────────────────────────────────────────────────────────────────────────────
def generate_excel(scan: dict, health: dict) -> bytes:
    if not _XLSX:
        raise RuntimeError("openpyxl not installed")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Drive Report"

    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(bold=True, color="FFFFFF")

    headers = ["Field", "Value"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font

    rows = [
        ("Drive",          scan.get("drive", "")),
        ("Model",          scan.get("model", "")),
        ("Serial",         scan.get("serial", "")),
        ("Type",           scan.get("drive_type", "")),
        ("Health Score",   health.get("score", "")),
        ("Grade",          health.get("grade", "")),
        ("Temperature °C", scan.get("smart", {}).get("temperature", "")),
        ("Power-On Hours", scan.get("smart", {}).get("power_on_hours", "")),
    ]
    for row_i, (k, v) in enumerate(rows, 2):
        ws.cell(row_i, 1, k).font = Font(bold=True)
        ws.cell(row_i, 2, v)

    ws2 = wb.create_sheet("SMART Attributes")
    hdrs = ["ID", "Name", "Value", "Raw"]
    for col, h in enumerate(hdrs, 1):
        cell = ws2.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
    for row_i, (aid, attr) in enumerate(scan.get("smart", {}).get("attributes", {}).items(), 2):
        ws2.cell(row_i, 1, aid)
        ws2.cell(row_i, 2, attr.get("name", ""))
        ws2.cell(row_i, 3, attr.get("value", ""))
        ws2.cell(row_i, 4, attr.get("raw", ""))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
#  JSON Export
# ─────────────────────────────────────────────────────────────────────────────
def generate_json(scan: dict, health: dict, risk: dict) -> bytes:
    payload = {
        "generated_at": datetime.datetime.now().isoformat(),
        "drive_info":   scan,
        "health":       health,
        "risk":         risk,
    }
    return json.dumps(payload, indent=2).encode("utf-8")
